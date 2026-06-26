import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Cores ────────────────────────────────────────────────────────────────────
COR_HEADER_GRUPO  = "1F3864"   # azul escuro — cabeçalho do grupo (BISMUT, GET…)
COR_HEADER_COL    = "2F5496"   # azul médio — cabeçalho das colunas (PARTE, COMPRAS…)
COR_HEADER_FONTE  = "FFFFFF"   # branco
COR_POSITIVO      = "C6EFCE"   # verde claro — NET positivo
COR_NEGATIVO      = "FFC7CE"   # vermelho claro — NET negativo
COR_ZERO          = "FFFFFF"   # branco — NET zero
COR_NET_VERDE_FNT = "006100"
COR_NET_VERM_FNT  = "9C0006"
COR_BORDA         = "9DC3E6"   # borda suave

FONT_HEADER = Font(name="Arial", bold=True, color=COR_HEADER_FONTE, size=9)
FONT_TITLE  = Font(name="Arial", bold=True, color=COR_HEADER_FONTE, size=10)
FONT_BODY   = Font(name="Arial", size=9)
FONT_TOTAL  = Font(name="Arial", bold=True, size=9)

FILL_GRUPO  = PatternFill("solid", fgColor=COR_HEADER_GRUPO)
FILL_COL    = PatternFill("solid", fgColor=COR_HEADER_COL)

ALN_CTR = Alignment(horizontal="center", vertical="center")
ALN_LFT = Alignment(horizontal="left",   vertical="center")
ALN_RGT = Alignment(horizontal="right",  vertical="center")

SIDE   = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)

FMT_NUM = '#,##0.000000'
FMT_MWH = '#,##0.000'


def fill_pos(val):
    if val is None or val == "":
        return PatternFill("solid", fgColor="FFFFFF")
    try:
        v = float(val)
    except:
        return PatternFill("solid", fgColor="FFFFFF")
    if v > 1e-9:
        return PatternFill("solid", fgColor=COR_POSITIVO)
    elif v < -1e-9:
        return PatternFill("solid", fgColor=COR_NEGATIVO)
    return PatternFill("solid", fgColor="FFFFFF")


def font_net(val):
    try:
        v = float(val)
    except:
        return FONT_BODY
    if v > 1e-9:
        return Font(name="Arial", size=9, color=COR_NET_VERDE_FNT)
    elif v < -1e-9:
        return Font(name="Arial", size=9, color=COR_NET_VERM_FNT)
    return Font(name="Arial", size=9)


def cell(ws, row, col, value=None, font=None, fill=None,
         alignment=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if alignment:
        c.alignment = alignment
    if border:
        c.border = border
    if number_format:
        c.number_format = number_format
    return c


def merge_header(ws, row, col_start, col_end, text, fill=FILL_GRUPO, font=FONT_TITLE):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start)
    c.value = text
    c.font = font
    c.fill = fill
    c.alignment = ALN_CTR
    c.border = BORDER


# ── Leitura dos dados ────────────────────────────────────────────────────────

def ler_cliq(sheet_name):
    df_raw = pd.read_excel(
        '/mnt/user-data/uploads/Book_Paloma.xlsm',
        sheet_name=sheet_name, header=0, dtype=str
    )
    df_raw.columns = df_raw.iloc[0]
    df = df_raw.drop(0).reset_index(drop=True)
    df['MWmedio'] = pd.to_numeric(df['MWmedio'], errors='coerce').fillna(0)
    return df


df_cliq = ler_cliq('CLIQCCEE')
df_bism = ler_cliq('CLIQCCEE_BISMUT')

df_conf = pd.read_excel(
    '/mnt/user-data/uploads/Book_Paloma.xlsm',
    sheet_name='Conferência', header=1, dtype=str
)
df_conf['Volume MWm']   = pd.to_numeric(df_conf['Volume MWm'],   errors='coerce').fillna(0)
df_conf['Volume (MWh)'] = pd.to_numeric(df_conf['Volume (MWh)'], errors='coerce').fillna(0)


# ── Definição dos grupos ─────────────────────────────────────────────────────

GRUPOS = [
    {
        'nome':        'BISMUT',
        'parte_book':  'NEWAVE BISMUT COMERCIALIZADORA DE ENERGIA S.A.',
        'df_cliq':     df_bism,
        'prefixo_cliq': 'BISMUT COM',
        # Siglas internas da parte (vendedor ou comprador deles mesmos)
        'siglas':      ['BISMUT COM', 'BISMUT COM I5', 'BISMUT COM I0',
                        'BISMUT COM I1', 'BISMUT COM CQ5'],
    },
    {
        'nome':        'GET',
        'parte_book':  'GET COMERCIALIZADORA DE ENERGIA S.A.',
        'df_cliq':     df_cliq,
        'prefixo_cliq': 'GET ENERGY TRADING',
        'siglas':      ['GET ENERGY TRADING', 'GET ENERGY TRADING I5',
                        'GET ENERGY TRADING I0', 'GET ENERGY TRADING I1',
                        'GET ENERGY TRADING CQ5'],
    },
    {
        'nome':        'MATRIX',
        'parte_book':  'MATRIX COMERCIALIZADORA DE ENERGIA ELETRICA S/A',
        'df_cliq':     df_cliq,
        'prefixo_cliq': 'MATRIX COM',
        'siglas':      ['MATRIX COM', 'MATRIX COM I5', 'MATRIX COM I0',
                        'MATRIX COM I1', 'MATRIX COM CQ5', 'MATRIX COM I8'],
    },
    {
        'nome':        'ARGENTUM',
        'parte_book':  'ARGENTUM COMERCIALIZADORA DE ENERGIA LTDA',
        'df_cliq':     df_cliq,
        'prefixo_cliq': 'ARGENTUM COM',
        'siglas':      ['ARGENTUM COM', 'ARGENTUM COM I5', 'ARGENTUM COM I0',
                        'ARGENTUM COM I1', 'ARGENTUM COM CQ5'],
    },
    {
        'nome':        'MATRIX CAMANDUCAIA',
        'parte_book':  'IBS SE-CAMANDUCAIA',
        'df_cliq':     df_cliq,
        'prefixo_cliq': 'MTX CAMANDUCAIA',
        'siglas':      ['MTX CAMANDUCAIA'],
    },
]


def calcular_grupo(grupo):
    """
    Retorna dict com:
      cliq_mwm:  { sigla: {compra, venda, net} }
      book_mwm:  { sigla: {compra, venda, net} }  — agrupado por perfil (Vendedor/Comprador)
      cliq_mwh:  idem em MWh
      book_mwh:  idem em MWh
    """
    df_c = grupo['df_cliq']
    pref = grupo['prefixo_cliq']
    parte = grupo['parte_book']

    # CLIQ ─ compras (sigla é o comprador) e vendas (sigla é o vendedor)
    mask_comp = df_c['SIGLA_PERFIL_COMPRADOR'].fillna('').str.startswith(pref)
    mask_vend = df_c['SIGLA_PERFIL_VENDEDOR'].fillna('').str.startswith(pref)

    cl_comp = df_c[mask_comp].groupby('SIGLA_PERFIL_COMPRADOR')['MWmedio'].sum()
    cl_vend = df_c[mask_vend].groupby('SIGLA_PERFIL_VENDEDOR')['MWmedio'].sum()

    # Conjunto de siglas encontradas
    all_siglas = sorted(set(cl_comp.index.tolist()) | set(cl_vend.index.tolist()))

    cliq_mwm = {}
    for s in all_siglas:
        c = float(cl_comp.get(s, 0))
        v = float(cl_vend.get(s, 0))
        cliq_mwm[s] = {'compra': c, 'venda': v, 'net': c - v}

    # BOOK ─ agrupado pela sigla (Comprador quando compra, Vendedor quando vende)
    book = df_conf[df_conf['Parte'] == parte].copy()
    book_comp = book[book['Operação'] == 'Compra'].groupby('Comprador')['Volume MWm'].sum()
    book_vend = book[book['Operação'] == 'Venda'].groupby('Vendedor')['Volume MWm'].sum()

    # Filtra só as siglas internas do grupo
    def filtrar(series):
        return {k: v for k, v in series.items()
                if str(k).startswith(pref)}

    bk_comp = filtrar(book_comp)
    bk_vend = filtrar(book_vend)
    all_book = sorted(set(bk_comp.keys()) | set(bk_vend.keys()))

    book_mwm = {}
    for s in all_book:
        c = float(bk_comp.get(s, 0))
        v = float(bk_vend.get(s, 0))
        book_mwm[s] = {'compra': c, 'venda': v, 'net': c - v}

    return cliq_mwm, book_mwm


# ── Criação da aba ───────────────────────────────────────────────────────────

wb = load_workbook('/mnt/user-data/uploads/Book_Paloma.xlsm', keep_vba=True)

# Remove aba Check se já existir
if 'Check' in wb.sheetnames:
    del wb['Check']

ws = wb.create_sheet('Check')

# Larguras das colunas
# Layout: [espaço] [col1:CLIQ CCEE label][col2:PARTE][col3:COMPRAS][col4:VENDAS][col5:NET]
#                  [gap]
#                  [col7:BOOK label][col8:PARTE][col9:COMPRAS][col10:VENDAS][col11:NET]
col_widths = {
    1:  1,   # margem
    2:  14,  # label bloco (CLIQ CCEE / BOOK)
    3:  22,  # PARTE
    4:  14,  # COMPRAS
    5:  14,  # VENDAS
    6:  14,  # NET
    7:  2,   # gap
    8:  14,  # label bloco (BOOK)
    9:  22,  # PARTE
    10: 14,  # COMPRAS
    11: 14,  # VENDAS
    12: 14,  # NET
}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# Linha de início
ROW = 2

COLS_CLIQ = (2, 3, 4, 5, 6)   # label, parte, compras, vendas, net
COLS_BOOK = (8, 9, 10, 11, 12)


def escrever_bloco(ws, row, grupo_nome, cliq_data, book_data, unidade='MWm'):
    """Escreve um bloco lado a lado (CLIQ | BOOK) e retorna a próxima linha."""

    # ── Cabeçalho do grupo ───────────────────────────────────────────────────
    merge_header(ws, row, COLS_CLIQ[0], COLS_CLIQ[4], grupo_nome)
    merge_header(ws, row, COLS_BOOK[0], COLS_BOOK[4], grupo_nome)
    row += 1

    # ── Cabeçalho das colunas ────────────────────────────────────────────────
    headers_cliq = ['CLIQ CCEE', 'PARTE', f'COMPRAS ({unidade})', f'VENDAS ({unidade})', f'NET ({unidade})']
    headers_book = ['BOOK',      'PARTE', f'COMPRAS ({unidade})', f'VENDAS ({unidade})', 'NET']

    for col, h in zip(COLS_CLIQ, headers_cliq):
        cell(ws, row, col, h, font=FONT_HEADER, fill=FILL_COL,
             alignment=ALN_CTR, border=BORDER)
    for col, h in zip(COLS_BOOK, headers_book):
        cell(ws, row, col, h, font=FONT_HEADER, fill=FILL_COL,
             alignment=ALN_CTR, border=BORDER)
    row += 1

    # ── Linhas de dados ──────────────────────────────────────────────────────
    # Unifica conjunto de siglas de ambos os blocos
    todas_siglas = sorted(set(list(cliq_data.keys()) + list(book_data.keys())))

    if not todas_siglas:
        # linha vazia para grupos sem dados
        for col in list(COLS_CLIQ) + list(COLS_BOOK):
            cell(ws, row, col, '-', font=FONT_BODY, alignment=ALN_CTR, border=BORDER)
        row += 1
        return row

    for sigla in todas_siglas:
        # CLIQ
        cd = cliq_data.get(sigla, {'compra': 0, 'venda': 0, 'net': 0})
        c_comp, c_vend, c_net = cd['compra'], cd['venda'], cd['net']

        # BOOK
        bd = book_data.get(sigla, {'compra': 0, 'venda': 0, 'net': 0})
        b_comp, b_vend, b_net = bd['compra'], bd['venda'], bd['net']

        fmt = FMT_NUM

        # Linha CLIQ
        cell(ws, row, COLS_CLIQ[0], '', fill=FILL_GRUPO, border=BORDER)
        cell(ws, row, COLS_CLIQ[1], sigla, font=FONT_BODY, alignment=ALN_LFT, border=BORDER)
        cell(ws, row, COLS_CLIQ[2], c_comp, font=FONT_BODY, alignment=ALN_RGT, border=BORDER, number_format=fmt)
        cell(ws, row, COLS_CLIQ[3], c_vend, font=FONT_BODY, alignment=ALN_RGT, border=BORDER, number_format=fmt)
        net_c = ws.cell(row=row, column=COLS_CLIQ[4])
        net_c.value = c_net
        net_c.font  = font_net(c_net)
        net_c.fill  = fill_pos(c_net)
        net_c.alignment = ALN_RGT
        net_c.border    = BORDER
        net_c.number_format = fmt

        # Linha BOOK
        cell(ws, row, COLS_BOOK[0], '', fill=FILL_GRUPO, border=BORDER)
        cell(ws, row, COLS_BOOK[1], sigla, font=FONT_BODY, alignment=ALN_LFT, border=BORDER)
        cell(ws, row, COLS_BOOK[2], b_comp, font=FONT_BODY, alignment=ALN_RGT, border=BORDER, number_format=fmt)
        cell(ws, row, COLS_BOOK[3], b_vend, font=FONT_BODY, alignment=ALN_RGT, border=BORDER, number_format=fmt)
        net_b = ws.cell(row=row, column=COLS_BOOK[4])
        net_b.value = b_net
        net_b.font  = font_net(b_net)
        net_b.fill  = fill_pos(b_net)
        net_b.alignment = ALN_RGT
        net_b.border    = BORDER
        net_b.number_format = fmt

        row += 1

    # ── Total ────────────────────────────────────────────────────────────────
    tot_c_comp = sum(d['compra'] for d in cliq_data.values())
    tot_c_vend = sum(d['venda']  for d in cliq_data.values())
    tot_c_net  = tot_c_comp - tot_c_vend
    tot_b_comp = sum(d['compra'] for d in book_data.values())
    tot_b_vend = sum(d['venda']  for d in book_data.values())
    tot_b_net  = tot_b_comp - tot_b_vend

    fill_tot = PatternFill("solid", fgColor="D9E1F2")

    for col in COLS_CLIQ:
        ws.cell(row=row, column=col).fill = fill_tot
        ws.cell(row=row, column=col).border = BORDER
    for col in COLS_BOOK:
        ws.cell(row=row, column=col).fill = fill_tot
        ws.cell(row=row, column=col).border = BORDER

    cell(ws, row, COLS_CLIQ[0], 'TOTAL', font=FONT_TOTAL, fill=fill_tot, alignment=ALN_CTR, border=BORDER)
    cell(ws, row, COLS_CLIQ[1], '', fill=fill_tot, border=BORDER)
    cell(ws, row, COLS_CLIQ[2], tot_c_comp, font=FONT_TOTAL, fill=fill_tot, alignment=ALN_RGT, border=BORDER, number_format=fmt)
    cell(ws, row, COLS_CLIQ[3], tot_c_vend, font=FONT_TOTAL, fill=fill_tot, alignment=ALN_RGT, border=BORDER, number_format=fmt)
    net_ct = ws.cell(row=row, column=COLS_CLIQ[4])
    net_ct.value = tot_c_net; net_ct.font = Font(name="Arial", bold=True, size=9)
    net_ct.fill = fill_tot; net_ct.alignment = ALN_RGT; net_ct.border = BORDER; net_ct.number_format = fmt

    cell(ws, row, COLS_BOOK[0], 'TOTAL', font=FONT_TOTAL, fill=fill_tot, alignment=ALN_CTR, border=BORDER)
    cell(ws, row, COLS_BOOK[1], '', fill=fill_tot, border=BORDER)
    cell(ws, row, COLS_BOOK[2], tot_b_comp, font=FONT_TOTAL, fill=fill_tot, alignment=ALN_RGT, border=BORDER, number_format=fmt)
    cell(ws, row, COLS_BOOK[3], tot_b_vend, font=FONT_TOTAL, fill=fill_tot, alignment=ALN_RGT, border=BORDER, number_format=fmt)
    net_bt = ws.cell(row=row, column=COLS_BOOK[4])
    net_bt.value = tot_b_net; net_bt.font = Font(name="Arial", bold=True, size=9)
    net_bt.fill = fill_tot; net_bt.alignment = ALN_RGT; net_bt.border = BORDER; net_bt.number_format = fmt

    row += 2  # espaço entre grupos
    return row


# ── Título da aba ────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 8

# ── Renderizar todos os grupos ───────────────────────────────────────────────
for grupo in GRUPOS:
    cliq_mwm, book_mwm = calcular_grupo(grupo)
    ROW = escrever_bloco(ws, ROW, grupo['nome'], cliq_mwm, book_mwm, 'MWm')

# ── Fixar painel (freeze) ─────────────────────────────────────────────────────
ws.freeze_panes = 'C3'

# ── Salvar ───────────────────────────────────────────────────────────────────
output_path = '/mnt/user-data/outputs/Book_Paloma_Check.xlsm'
wb.save(output_path)
print(f'Salvo em: {output_path}')
